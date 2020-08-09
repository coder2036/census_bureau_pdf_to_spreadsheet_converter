# -*- coding: utf-8 -*-
"""
Created on Tue Jun 16 10:10:40 2020

@author: gupta018
"""

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment,Font
from openpyxl.styles.borders import Border,Side
from openpyxl.styles.colors import Color
from openpyxl.styles.fills import PatternFill

def write_sheet(df,output_file_path,company_name):

    
    # fix the position of the confidentiality thing
    
    wb = Workbook()
    ws = wb['Sheet']
    ws.title = '2020 Reporting Calendar'
    ws['A1'] = company_name + '\n2020 Reporting Calendar'
    font = Font(size=20,bold=True)
    ws['A1'].font = font
    color = Color('FFC000')
    fill = PatternFill(patternType='solid',fgColor=color)
    ws['A1'].fill = fill
    ws.merge_cells('A1:J1')
    ws.row_dimensions[1].height = 60
     
    for row in dataframe_to_rows(df,index=False):
        ws.append(row)
        
    for row in ws.iter_cols(min_row=3,min_col=10,max_col=10):
        for cell in row:
            print('hit')
            print(cell.value)
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"
    
    curr_survey = None
    num_rows = 0
    start_row = 3
    cols_to_merge = [1,2,3,4,5,8,9,10,11]
    for index,row in df.iterrows():
        print('start_row: ' + str(start_row) + ', num_rows: ' + str(num_rows))
        if row[0] != curr_survey:
            if num_rows > 1:
                for col in cols_to_merge:
                    ws.merge_cells(start_row=start_row,end_row=start_row+num_rows-1,start_column=col,end_column=col)
            curr_survey = row[0]
            start_row += num_rows
            num_rows = 1
        else:
            num_rows += 1        
    if num_rows > 1:
        for col in cols_to_merge:
            ws.merge_cells(start_row=start_row,end_row=start_row+num_rows-1,start_column=col,end_column=col)
        
    ## confidentiality warning
    warning_row = df.shape[0] + 1 + 1 + 2 # +1 for title and for headers. 
                                            # +2 for how far below everything else
    warning_col_begin = 'D'
    warning_col_end = 'J'
    warning_loc_begin = warning_col_begin + str(warning_row)
    warning_loc_end = warning_col_end + str(warning_row)
    ws[warning_loc_begin] = 'Disclosure Prohibited Title 13/26 U.S.C. – Census Confidential'
    ws.merge_cells(warning_loc_begin + ':' + warning_loc_end)
    ws[warning_loc_begin].fill = PatternFill(patternType='solid',fgColor=Color('BF8F00'))
    ws[warning_loc_begin].font = Font(bold=True)
    
    ## adding green background to columns 11 and 12 from row 5 onwards
    green = PatternFill(patternType='solid',fgColor=Color('E2EFDA'))
    for col in ws.iter_cols(min_col=11,max_col=12,min_row=3,max_row=3 + df.shape[0] - 1):
        for cell in col:
            cell.fill = green 
    
    # center all cells
    for row in ws.iter_rows():  
        for cell in row:      
            cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='top')
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 19
    ws.column_dimensions['E'].width = 19
    ws.column_dimensions['F'].width = 19
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 22
    ws.column_dimensions['I'].width = 23
    ws.column_dimensions['J'].width = 21
    ws.column_dimensions['K'].width = 17
    ws.column_dimensions['L'].width = 26
        
    

    
    # bolding mandatory/volunatry
    for row in ws.iter_cols(min_row=3,min_col=2,max_col=2):
        for cell in row:
            cell.font = Font(bold=True)
            
    for col in ws.iter_cols(min_col=1,max_col=10,min_row=2,max_row=2):
        for cell in col:
            cell.fill = PatternFill(patternType='solid',fgColor=Color('fff2cc'))
            
    for col in ws.iter_cols(min_col=11,max_col=12,min_row=2,max_row=2):
        for cell in col:           
            cell.fill = PatternFill(patternType='solid',fgColor=Color('dbdbdb'))
            
    row = ws.row_dimensions[2]
    for row in ws.iter_rows(min_row=2,max_row=2):
        for cell in row:
            cell.font = Font(bold=True)
    
    for row in ws.iter_rows(min_row=2,max_row=2 + df.shape[0] + 1 + 1 - 2,min_col=1,max_col=12):
        for cell in row:
            cell.border = Border(left=Side(style='thin'),right=Side(style='thin'),
                                        top=Side(style='thin'),bottom=Side(style='thin'))
    
    wb.save(output_file_path)
    
    
#def test_formatting():
   # import pandas as pd
   # write_sheet(path,pd.DataFrame(),'Random Number',"Company Name")
    
#path = r'H:/pdf_to_xlsx'

#write_sheet(path,None,'ID######','Company')
#test_formatting()